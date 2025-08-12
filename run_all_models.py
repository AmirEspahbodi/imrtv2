import argparse
import random
import os
import json
from pathlib import Path
from typing import List, Dict, Any

import openpyxl
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
import subprocess
import sys


def load_workbook(path: Path) -> Workbook:
    print(f"Loading workbook from {path}")
    return openpyxl.load_workbook(path)


def get_headers(sheet: Worksheet) -> List[str]:
    headers = [cell.value for cell in sheet[1]]
    print(f"Headers found: {headers}")
    return headers  # type: ignore


def process_rows(
    sheet: Worksheet,
    headers: List[str],
    workbook: Workbook,
    path: Path,
) -> None:
    acc_index = headers.index("acc")
    metrics = ["acc", "f1", "auc", "precision", "recall"]
    metrics_indexes = {i: headers.index(i) for i in metrics}
    resutl_path = "resutl_path"
    resutl_path_index = headers.index(resutl_path)
    # TIP
    # sync with hydra config
    save_result = "/home/amirh/work/medical_image_classification/result"

    for row in sheet.iter_rows(min_row=2):
        print(
            "\n+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++"
        )
        print(row[0].row, end="\n\n\n")

        # Map row values to header keys
        row_data: Dict[str, Any] = {
            header: row[idx].value for idx, header in enumerate(headers)
        }
        model = row_data["model"]

        if isinstance(row_data["backbone_trainable_layers"], str):
            btl = [
                str(i) for i in row_data["backbone_trainable_layers"].strip().split(" ")
            ]
        elif isinstance(row_data["backbone_trainable_layers"], int):
            btl = [str(row_data["backbone_trainable_layers"])]
        else:
            btl = [""]

        if isinstance(row_data["vit1_feature_strame"], str):
            v1fs = [str(i) for i in row_data["vit1_feature_strame"].strip().split(" ")]
        elif isinstance(row_data["vit1_feature_strame"], int):
            v1fs = [str(row_data["vit1_feature_strame"])]
        else:
            raise RuntimeError()

        if isinstance(row_data["vit2_feature_strame"], str):
            v2fs = [str(i) for i in row_data["vit2_feature_strame"].strip().split(" ")]
        elif isinstance(row_data["vit2_feature_strame"], int):
            v2fs = [str(row_data["vit2_feature_strame"])]
        else:
            raise RuntimeError()
        tp = row_data["training_plan"].strip()

        model_id = f"btl{''.join(btl)}_v1fs{''.join(v1fs)}_v2fs{''.join(v2fs)}_tp{tp}"
        args = f" dataset=chest_X_ray network={model} --btl {' '.join(btl)} --v1fs {' '.join(v1fs)} --v2fs {' '.join(v2fs)} --tp {tp}"

        resutl_path_value = row[resutl_path_index].value
        if resutl_path_value is not None:
            print(
                f"{row[0].row}==> this model is calculated {model}_{model_id} resutl path is {resutl_path_value} <=="
            )
            continue

        # print(" >> start check <<")
        # print(f"    row_data['model'] = {row_data['model']}, type={type(row_data['model'])}")
        # print(f"    row_data['backbone_trainable_layers'] = {row_data['backbone_trainable_layers']} , type={type(row_data['backbone_trainable_layers'])}")
        # print(f"    row_data['vit1_feature_strame'] = {row_data['vit1_feature_strame']} , type={type(row_data['vit1_feature_strame'])}")
        # print(f"    row_data['vit2_feature_strame'] = {row_data['vit2_feature_strame']} , type={type(row_data['vit2_feature_strame'])}")
        # print(f"    model = {model}")
        # print(f"    btl = {btl}")
        # print(f"    v1fs = {v1fs}")
        # print(f"    v2fs = {v2fs}")
        print(f"    model_id = {model_id}")
        print(f"    args = {args}")
        # print(" >> end check <<")

        ret = os.system(f"{sys.executable} main.py {args}")
        if os.name != "nt":
            if os.WIFEXITED(ret):
                exit_code = os.WEXITSTATUS(ret)
            else:
                # killed by signal
                sig = ret & 0xFF
                print(f"main_a.py was terminated by signal {sig!r}")
                sys.exit(1)
        else:
            # on Windows, os.system() just returns the process exit code
            exit_code = ret

        # now check
        if exit_code == 0:
            result_directory = f"{save_result}/{model}_{model_id}"
            finall_resutl_path = f"{result_directory}/final_weights_results.json"
            validation_resutl_path = (
                f"{result_directory}/best_validation_weights_results.json"
            )
            with open(finall_resutl_path, "r", encoding="utf-8") as f:
                finall_resutl = json.load(f)
            with open(validation_resutl_path, "r", encoding="utf-8") as f:
                validation_resutl = json.load(f)
            result = None
            if finall_resutl["acc"] < validation_resutl["acc"]:
                result = validation_resutl
            else:
                result = finall_resutl
            for metric, value in result.items():
                row[metrics_indexes[metric]].value = float(value)

            row[resutl_path_index].value = f"{save_result}/{model}/{model_id}"
        else:
            for metric in metrics:
                row[metrics_indexes[metric]].value = "FAILD"
                row[resutl_path_index].value

        workbook.save(path)
        print(f"Workbook saved after updating row {row[0].row}")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Process an Excel file: assign random accuracy per row and save changes immediately."
    )
    parser.add_argument(
        "input", type=Path, help="Path to the Excel file (.xlsx) to process"
    )
    args = parser.parse_args()
    input_path = args.input

    wb = load_workbook(input_path)
    ws = wb.active

    headers = get_headers(ws)
    if "acc" not in headers:
        print("Header 'acc' not found in the Excel sheet.")
        return

    process_rows(ws, headers, wb, input_path)


if __name__ == "__main__":
    main()
